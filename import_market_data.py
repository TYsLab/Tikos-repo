import os
import pyodbc
import openpyxl
from dotenv import load_dotenv

load_dotenv()

DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

if not DB_USER or not DB_PASSWORD:
    raise ValueError("DB_USER and DB_PASSWORD must be set in your .env file")

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "market_data.xlsx")

def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Row 3 has the dates (Jan 22, Jan 29, ...)
    # Row 1-2 are headers, data starts row 4
    date_row = [ws.cell(3, c).value for c in range(3, ws.max_column + 1)]

    rows = []
    current_category = None

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not any(row):
            continue
        if row[0]:  # new category
            current_category = row[0]
        asset = row[1]
        if not asset or "Notes" in str(asset):
            continue

        for i, date_label in enumerate(date_row):
            if date_label is None:
                continue
            close_price = row[i + 2]  # offset by 2 (Category, Asset cols)
            if close_price is None or str(close_price).startswith("="):
                continue
            try:
                close_price = float(close_price)
            except (ValueError, TypeError):
                continue

            rows.append({
                "AssetDate": str(date_label),
                "Category":  current_category,
                "Asset":     asset,
                "ClosePrice": close_price,
            })

    return rows

def import_to_sql(rows):
    print(f"Connecting to Azure SQL...")
    conn = pyodbc.connect(
        "Driver={ODBC Driver 18 for SQL Server};"
        "Server=tcp:tikos-sql-server.database.windows.net,1433;"
        "Database=tikos-market-db;"
        f"UID={DB_USER};"
        f"PWD={DB_PASSWORD};"
        "Encrypt=yes;"
        "TrustServerCertificate=no;"
    )
    cursor = conn.cursor()

    # Drop and recreate table to ensure correct schema
    cursor.execute("IF OBJECT_ID('MarketData', 'U') IS NOT NULL DROP TABLE MarketData")
    cursor.execute("""
        CREATE TABLE MarketData (
            Id          INT IDENTITY(1,1) PRIMARY KEY,
            AssetDate   NVARCHAR(50),
            Category    NVARCHAR(100),
            Asset       NVARCHAR(100),
            ClosePrice  FLOAT
        )
    """)
    conn.commit()

    inserted = 0
    for r in rows:
        cursor.execute("""
            INSERT INTO MarketData (AssetDate, Category, Asset, ClosePrice)
            VALUES (?, ?, ?, ?)
        """, r["AssetDate"], r["Category"], r["Asset"], r["ClosePrice"])
        inserted += 1

    conn.commit()
    conn.close()
    print(f"Done — {inserted} rows inserted into MarketData.")

if __name__ == "__main__":
    rows = parse_excel()
    print(f"Parsed {len(rows)} rows from {EXCEL_FILE}")
    import_to_sql(rows)
